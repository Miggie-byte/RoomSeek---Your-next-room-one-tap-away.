from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import boto3
import tempfile

dynamodb = boto3.resource('dynamodb', region_name='ap-southeast-1')
table = dynamodb.Table('rooms')
s3 = boto3.client('s3')

def lambda_handler(event, context):
    # Get the uploaded file from S3
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = event['Records'][0]['s3']['object']['key']

    # Download to temp file
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        s3.download_file(bucket, key, tmp.name)
        workbook = load_workbook(tmp.name)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        room_number = str(worksheet['A6'].value)
        output = { room_number: {} }

        for col in range(2, 14, 2):
            vacant_start = None
            vacant_end = None
            header_value = worksheet[f'{get_column_letter(col)}7'].value

            output[room_number][header_value] = []

            for row in range(8, 92):
                cell = worksheet.cell(row=row, column=col)
                is_vacant = cell.fill.fill_type is None

                if is_vacant:
                    if vacant_start is None:
                        vacant_start = row
                    vacant_end = row
                else:
                    if vacant_start is not None:
                        start_time = worksheet[f'A{vacant_start}'].value
                        end_time = worksheet[f'A{vacant_end}'].value
                        output[room_number][header_value].append({
                            "start_time": start_time,
                            "end_time": end_time
                        })
                        vacant_start = None

            if vacant_start is not None:
                start_time = worksheet[f'A{vacant_start}'].value
                end_time = worksheet[f'A{vacant_end}'].value
                output[room_number][header_value].append({
                    "start_time": start_time,
                    "end_time": end_time
                })

        table.put_item(Item={
            'room_number': room_number,
            'vacant_slots': output[room_number]
        })
        print(f"Stored {room_number} to DynamoDB")

    return {'statusCode': 200, 'body': 'Done'}
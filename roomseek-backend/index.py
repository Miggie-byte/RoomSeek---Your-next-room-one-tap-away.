from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

workbook = load_workbook('openpyxl-grades.xlsx')
worksheet = workbook.active
print(worksheet['A1'].value) # Print the value of cell A1

#Changing the value of cell A1
worksheet['A1'] = 'First Name'
print(worksheet['A1'].value) # Print the value of updated cell A1

#Saving Workbooks (All changes will not be applied if not saved)
#workbook.save('openpyxl-grades.xlsx') # Save the workbook with the updated value

#Creating, Listing and Changing Sheets
workbook.create_sheet('New Sheet') # Create a new sheet named Sheet3
print(workbook.sheetnames) # Print the names of all sheets in the workbook

#Accessing other sheets in the workbook
worksheet2 = workbook['Sheet2'] # Accessing Sheet2
print(worksheet2['A1'].value) # Print the value of cell A1 in Sheet2

#Creating a new workbook and adding data to it
workbook2 = load_workbook("room_schedule.xlsx") # Create a new workbook
worksheet3 = workbook2.active # Get the active sheet in the new workbook
#worksheet3.title = 'Room Schedule' # Rename the active sheet to 'Room Schedule'
"""
worksheet3.append(['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']) # Add a header row to the sheet

worksheet3.append(['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']) # Add a header row to the sheet

worksheet3.append(['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']) # Add a header row to the sheet

worksheet3.append(['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']) # Add a header row to the sheet

worksheet3.append(['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']) # Add a header row to the sheet

worksheet3.append(['end']) # Add a header row to the sheet

workbook2.save('room_schedule.xlsx') # Save the new workbook with the name 'room_schedule.xlsx'
"""
#Accessing Multiple Cells
for row in range(1,6):  
    for col in range(1,6):
        char = get_column_letter(col) # Get the column letter for the current column index
        cell_value = worksheet3[f'{char}{row}'].value # Get the value of the current cell
        print(cell_value, end=' ') # Print the value of the current cell
    print() # Print a newline after each row

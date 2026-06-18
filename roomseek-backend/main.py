from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter

workbook = load_workbook('Room Schedule.xlsx')
worksheet = workbook['1CSA']

room_number = worksheet['A6'].value
print("Room Number: "+str(room_number)) # Print the value of cell A6

#Looping through the cells and checking for vacancy
for col in range(2,14,2):  
    vacant_start = None # Variable to track the start of a vacant period
    vacant_end = None # Variable to track the end of a vacant period
    header_value = worksheet[f'{get_column_letter(col)}7'].value # Get the header value for the current column

    for row in range(8,92):
        cell = worksheet.cell(row=row, column=col) # Get the current cell
        is_vacant = cell.fill.fill_type is None # Check if the cell is vacant (no fill)

        if is_vacant:
            if vacant_start is None: # Start of a vacant period
                vacant_start = row
            vacant_end = row # Update the end of the vacant period

        else:
            if vacant_start is not None: # End of a vacant period
                char = 'A' # Get the column letter for the current column index
                start_time = worksheet[f'{char}{vacant_start}'].value # Get the start time of the vacant period
                end_time = worksheet[f'{char}{vacant_end}'].value # Get the end time of the vacant period
                print(f"{room_number} is vacant from {start_time} to {end_time} on {header_value}") # Print the vacant period for the room
                vacant_start = None # Reset the start of the vacant period

    # After the row loop, handle interval that ends at the last row
    if vacant_start is not None:
        start_time = worksheet[f'A{vacant_start}'].value
        end_time = worksheet[f'A{vacant_end}'].value
        print(f"{room_number} is vacant from {start_time} to {end_time} on {header_value}")